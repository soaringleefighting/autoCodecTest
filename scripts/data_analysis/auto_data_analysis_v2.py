#_*_ coding=utf-8_*_  #脚本中有中文注释必须包含这一句

#######################################################################################
##脚本功能： 本脚本用于计算BD-rate，Delta_PSNR, Delta_time，并绘制率失真曲线图。
##脚本用法： python auto_data_analysis.py anchor_result ./out refer1_result [refer2_result refer3_result]
##参数说明：    anchor_result     :	  原始测试数据
##             refer_result      :   对比测试数据
##             ./out             :   输出结果目录
##
## Created by lipeng at July 10 2020
## Version 2.2
## Modified:
## 2020.7.10 create tag v1.0
## 2020.7.16 create tag v2.0 support BDBR collect
## 2021.7.30 create tag v2.1 support calculate BDBR and plot RD curve
## 2021.7.31 create tag v2.2 支持三路或四路编码器对比计算BD-rate和绘制率失真曲线图(绘制在一张图上)
#######################################################################################
import os
import re
import sys
import glob
import filecmp
import shutil
import subprocess
import subprocess as sub
import csv
import codecs
from   collections import OrderedDict
import collections
import xlrd
import xlwt
from   xlutils.copy import copy
#import scipy 							#特殊三对角矩阵求解
import numpy as np 						#用于数学计算(曲线拟合、积分等)
import matplotlib.pyplot as plt         #绘制率失真曲线
import pandas as pd                     #转换excel
import openpyxl                         #Excel文件
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl import load_workbook

space     = ' '
delimiter = '/'

reload(sys)
sys.setdefaultencoding('utf-8')

#比较两个文件是否相同, 相同则返回True, 不同返回False
def	yuv_cmp(file1,file2):
	isNul1 = os.path.getsize(file1)
	isNul2 = os.path.getsize(file2)
	if((not isNul1) or (not isNul2)):
		return False
	if(isNul1 == isNul2):
		return True

#提取文件的名字
def get_file_name(fullfilename):
	tmp = fullfilename.strip()
	name = os.path.split(tmp)[-1]   #提取文件名，不包含路径
	return os.path.splitext(name)[0] #提取文件名，不包含后缀

#创建文件目录
def make_all_dir(path):	
	path = path.strip() #去除首位空格
	path=path.rstrip("\\")  #去除尾部\符号

	isExist = os.path.exists(path) #判断路径是否存在
        if not isExist:        #如果不存在则创建目录
                os.makedirs(path)
                print '[info]: ' + path+' Create success!'
                return True
        else:   #如果目录存在则不创建，并提示目录已经存在
                print '[info]: ' + path+' aleady exist!'
                return False

#获取数据文件
def	get_raw_data(rawdir):
	isfile = 0
	if os.path.isdir(rawdir):
		allfiles = os.listdir(rawdir)
		files = [rawdir+"/"+f for f in allfiles if re.search('txt$',f) or  re.search('log$',f)]
	elif os.path.isfile(rawdir):
		isfile = 1
		files = [rawdir]
	else:
		files = []
		print("ERROR: " + sys.argv[1] + "  is not a dir or file!")
	files.sort(key=str.lower)
	return [files,isfile]

#create new excel file
def create_excel(excel_name):
    file = open(excel_name,'wb')
    writer=csv.writer(file)
    file.close()


#collect data from format text to BDBR excel for ref data
def collect_data_to_BDBRexcel(exceldata, datawt, inputfile, outexcel):
    pFile = open(inputfile, 'a+')
    lines = pFile.readlines()
    data = collections.OrderedDict()  ##使用有序字典
    splitValue = lines[0].strip().split(' ')
    data[splitValue[0]] = [splitValue[0], splitValue[1], splitValue[2], splitValue[3]]
    sequence_name_plus_qp = splitValue[0]
    sequence_name = splitValue[0].split('_qp')[0]
    sequence_qp = splitValue[0].split('_qp')[1]

    table = exceldata.sheet_by_name('AI-Main')
    table_wt = datawt.get_sheet('AI-Main')

    ##遍历excel中每一行，存在匹配的字符串则写入对应的bitrate,Y-PSNR和EncT
    nrows = table.nrows
    for i in range(nrows):
        if type(table.col_values(2)[i]) == float:  ##将float类型转换成int类型
            qp = int(table.col_values(2)[i])
        if str(table.col_values(1)[i]) == sequence_name and str(qp) == sequence_qp:
            table_wt.write(i, 11, splitValue[1]) #write bitrate
            table_wt.write(i, 12, splitValue[2]) #write Y-PSNR
            table_wt.write(i, 15, splitValue[3]) #write EncT(s)
    datawt.save(outexcel)
    return 0


# collect data from format text to BDBR excel for anchor data
def collect_data_to_BDBRexcel_vs(exceldata, datawt, inputfile, outexcel):
    pFile = open(inputfile, 'a+')
    lines = pFile.readlines()
    data = collections.OrderedDict()  ##使用有序字典
    splitValue = lines[0].strip().split(' ')
    data[splitValue[0]] = [splitValue[0], splitValue[1], splitValue[2], splitValue[3]]
    sequence_name_plus_qp = splitValue[0]
    sequence_name = splitValue[0].split('_qp')[0]
    sequence_qp = splitValue[0].split('_qp')[1]

    table = exceldata.sheet_by_name('AI-Main')
    table_wt = datawt.get_sheet('AI-Main')

    ##遍历excel中每一行，存在匹配的字符串则写入对应的bitrate,Y-PSNR和EncT
    nrows = table.nrows
    for i in range(nrows):
        if type(table.col_values(2)[i]) == float:  ##将float类型转换成int类型
            qp = int(table.col_values(2)[i])
        if str(table.col_values(1)[i]) == sequence_name and str(qp) == sequence_qp:
            table_wt.write(i, 3, splitValue[1]) #write bitrate
            table_wt.write(i, 4, splitValue[2]) #write Y-PSNR
            table_wt.write(i, 7, splitValue[3]) #write EncT(s)
            ## 计算编码节省时间
    datawt.save(outexcel)
    return 0

## 从特定格式csv文件中提取码率，PSNR和time信息
def shuffle_info(anchor, isAnchor=1, refer_idx=0):
    anchordata    = csv.reader(open(anchor), quotechar="'") ## 'r'
    count_num = -1
    index_num_anchor = 0
    index_num_ref    = 0

    origBit_arr    = np.zeros(4)
    testBit_arr    = np.zeros(4)
    origYPSNR_arr  = np.zeros(4)
    testYPSNR_arr  = np.zeros(4)
    origUPSNR_arr  = np.zeros(4)
    testUPSNR_arr  = np.zeros(4)
    origVPSNR_arr  = np.zeros(4)
    testVPSNR_arr  = np.zeros(4)
    origTime_arr   = np.zeros(4)
    testTime_arr   = np.zeros(4)

    #print total_index
    for anchor_line in anchordata:
        count_num = count_num + 1
        if count_num == 0:
            continue
        #seq_name=anchor_line[0].split('_br')[0]  # just for vbr mode
        seq_name = '_'.join([anchor_line[0].split('_')[0], anchor_line[0].split('_')[1]])
        #print seq_name
        bitrate = anchor_line[2]
        Y_PSNR  = anchor_line[3]
        U_PSNR  = anchor_line[4]
        V_PSNR  = anchor_line[5]
        time    = anchor_line[-1]
    
        if count_num%2 == 1:
            origBit_arr[index_num_anchor]    = bitrate
            origYPSNR_arr[index_num_anchor]  = Y_PSNR
            origUPSNR_arr[index_num_anchor]  = U_PSNR
            origVPSNR_arr[index_num_anchor]  = V_PSNR
            origTime_arr[index_num_anchor]   = time
            index_num_anchor=index_num_anchor+1
            #print 'index_num_anchor:', index_num_anchor
                
        else:
            testBit_arr[index_num_ref]    = bitrate
            testYPSNR_arr[index_num_ref]  = Y_PSNR
            testUPSNR_arr[index_num_ref]  = U_PSNR
            testVPSNR_arr[index_num_ref]  = V_PSNR
            testTime_arr[index_num_ref]   = time
            index_num_ref=index_num_ref+1
            #print 'index_num_ref:', index_num_ref   

        if (index_num_anchor == 4):
            index_num_anchor = 0 
        if (index_num_ref == 4):
            index_num_ref = 0
            #print 'count_num:',count_num
            if isAnchor == 1:
                #print seq_name, oriBit_arr
                origBit_dict[seq_name]  = origBit_arr.copy()  #深拷贝
                origYPSNR_dict[seq_name] = origYPSNR_arr.copy()
                origUPSNR_dict[seq_name] = origUPSNR_arr.copy()
                origVPSNR_dict[seq_name] = origVPSNR_arr.copy()
                origTime_dict[seq_name] = origTime_arr.copy()
                seqName_dict[count_num/8] = seq_name
                #print origBit_dict
                #print origYPSNR_dict
                #print origTime_dict
                #print seqName_dict

                testBit_dict  [seq_name] = testBit_arr.copy()
                testYPSNR_dict[seq_name] = testYPSNR_arr.copy()
                testUPSNR_dict[seq_name] = testUPSNR_arr.copy()
                testVPSNR_dict[seq_name] = testVPSNR_arr.copy()
                testTime_dict [seq_name] = testTime_arr.copy()
                #print testBit_dict
                #print testYPSNR_dict
                #print testTime_dict
    return (count_num/8)

def csv_to_xlsx(csvfile, excelfile):
    with open(csvfile, 'r') as f:
        read = csv.reader(codecs.EncodedFile(f, 'utf8', 'utf_8_sig'), delimiter=" ")
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('result')  # 创建一个sheet表格
        l = 0
        for line in read:
            #print(line)
            r = 0
            for i in line:
                #print(i)
                sheet.write(l, r, unicode(i, 'utf-8'))  # 一个一个将单元格数据写入
                r = r + 1
            l = l + 1
        workbook.save(excelfile)  # 保存Excel

def pchip_end(h1, h2, del1, del2):
    d = ((2*h1 + h2)*del1 - h1*del2) / (h1 + h2)
    if np.sign(d) != np.sign(del1):
        d = 0
    elif np.sign(del1) != np.sign(del2) and np.abs(d) > np.abs(3*del1):
        d = 3 * del1
    return d

def pchip_slopes(h, delta):
    d = np.zeros(len(h) + 1)
    k = np.argwhere(np.sign(delta[:-1]) * np.sign(delta[1:]) > 0).reshape(-1) + 1
    w1 = 2*h[k] + h[k-1]
    w2 = h[k] + 2*h[k-1]
    d[k] = (w1 + w2) / (w1 / delta[k-1] + w2 / delta[k])
    d[0] = pchip_end(h[0], h[1], delta[0], delta[1])
    d[-1] = pchip_end(h[-1], h[-2], delta[-1], delta[-2])
    return d

def spline_slopes(h, delta):
    a, r = np.zeros([3, len(h)+1]), np.zeros(len(h)+1)
    a[0, 1], a[0, 2:] = h[0] + h[1], h[:-1]
    a[1, 0], a[1, 1:-1], a[1, -1] = h[1], 2*(h[:-1] + h[1:]), h[-2]
    a[2, :-2], a[2, -2] = h[1:], h[-1] + h[-2]

    r[0] = ((h[0] + 2*a[0, 1])*h[1]*delta[0] + h[0]**2*delta[1]) / a[0, 1]
    r[1:-1] = 3*(h[1:] * delta[:-1] + h[:-1] * delta[1:])
    r[-1] = (h[-1]**2*delta[-2] + (2*a[2, -2] + h[-1])*h[-2]*delta[-1]) / a[2, -2]

    d = scipy.linalg.solve_banded((1, 1), a, r)
    return d

class PCHIP:
    def __init__(self, x, y, use_spline=False):
        assert len(np.unique(x)) == len(x)
        #将数据按x坐标从小到大排序
        order = np.argsort(x)
        self.xi, self.yi = x[order], y[order]

        #求输入x、y的间隔，按间隔计算delta
        h = np.diff(self.xi)
        delta = np.diff(self.yi) / h

        self.d = spline_slopes(h, delta) if use_spline else pchip_slopes(h, delta)
        self.c = (3*delta - 2*self.d[:-1] - self.d[1:]) / h
        self.b = (self.d[:-1] - 2*delta + self.d[1:]) / h**2

        """
        The piecewise function is like p(x) = y_k + s*d_k + s*s*c_k + s*s*s*b_k
        where s = x - xi_k, k is the interval includeing x.
        So the original function of p(x) is P(x) = s*y_k + 1/2*s*s*d_k + 1/3*s*s*s*c_k + 1/4*s*s*s*s*b_k + C.
        """
        self.interval_int_coeff = []
        self.interval_int = np.zeros(len(x)-1)
        for i in range(len(x)-1):
            self.interval_int_coeff.append(np.polyint([self.b[i], self.c[i], self.d[i], self.yi[i]]))
            self.interval_int[i] = np.polyval(self.interval_int_coeff[-1], h[i]) - np.polyval(self.interval_int_coeff[-1], 0)

    def _integral(self, lower, upper):
        assert lower <= upper
        if lower < np.min(self.xi):
            lower = np.min(self.xi)
            print('Warning: The lower bound is less than the interval and clipped!')
        elif lower > np.max(self.xi):
            print('Warning: The lower bound is greater than the interval!')
            return 0
        if upper > np.max(self.xi):
            upper = np.max(self.xi)
            print('Warning: The upper bound is greater than the interval and clipped!')
        elif upper < np.min(self.xi):
            print('Warning: The lower bound is less than the interval!')
            return 0
        left = np.arange(len(self.xi))[self.xi - lower > -1e-6][0]
        right = np.arange(len(self.xi))[self.xi - upper < 1e-6][-1]

        inte = np.sum(self.interval_int[left:right])
        if self.xi[left] - lower > 1e-6:
            inte += (np.polyval(self.interval_int_coeff[left-1], self.xi[left]-self.xi[left-1]) - np.polyval(self.interval_int_coeff[left-1], lower-self.xi[left-1]))
        if self.xi[right] - upper < -1e-6:
            inte += (np.polyval(self.interval_int_coeff[right], upper-self.xi[right]) - np.polyval(self.interval_int_coeff[right], 0))
        return inte

    def integral(self, lower, upper):
        if lower > upper:
            return -self._integral(upper, lower)
        else:
            return self._integral(lower, upper)

def computeBDRate(testNum, basePSNR, baseBitrate, testPSNR, testBitrate, piecewise=True):
    #码率取对数
    baseLogBitrate = np.log10(baseBitrate)
    testLogBitrate = np.log10(testBitrate)
    #确定共同范围
    minPSNR = np.max((np.min(basePSNR), np.min(testPSNR)))
    maxPSNR = np.min((np.max(basePSNR), np.max(testPSNR)))
    if piecewise == True:
        baseIntegral = PCHIP(basePSNR, baseLogBitrate, use_spline=False).integral(minPSNR, maxPSNR)
        testIntegral = PCHIP(testPSNR, testLogBitrate, use_spline=False).integral(minPSNR, maxPSNR)
    else:
        #拟合曲线
        baseFitting = np.polyfit(basePSNR, baseLogBitrate, testNum-1)
        testFitting = np.polyfit(testPSNR, testLogBitrate, testNum-1)
        #不定积分
        baseIndefiniteInt = np.polyint(baseFitting)
        testIndefiniteInt = np.polyint(testFitting)
        #求积分
        baseIntegral = np.polyval(baseIndefiniteInt, maxPSNR) - np.polyval(baseIndefiniteInt, minPSNR)
        testIntegral = np.polyval(testIndefiniteInt, maxPSNR) - np.polyval(testIndefiniteInt, minPSNR)

    #平均差值
    meanEXPDiff = (testIntegral - baseIntegral) / (maxPSNR - minPSNR)
    meanDiff = (np.power(10, meanEXPDiff) - 1)

    return meanDiff


def ComputeBDBR_PSNR_Time(index_num):
    Delta_YPSNR        = 0.0
    Delta_UPSNR        = 0.0
    Delta_VPSNR        = 0.0
    Delta_time         = 0.0
    Delta_YPSNR_2      = 0.0
    Delta_UPSNR_2      = 0.0
    Delta_VPSNR_2      = 0.0
    Delta_time_2       = 0.0
    Delta_YPSNR_3      = 0.0
    Delta_UPSNR_3      = 0.0
    Delta_VPSNR_3      = 0.0
    Delta_time_3       = 0.0    
    
    global BDBRP_avg    
    global BDBR_avg       
    global Delta_YPSNR_avg 
    global Delta_UPSNR_avg 
    global Delta_VPSNR_avg
    global Delta_time_avg 
    global BDBRP_avg_2    
    global BDBR_avg_2  
    global Delta_YPSNR_avg_2
    global Delta_UPSNR_avg_2
    global Delta_VPSNR_avg_2 
    global Delta_time_avg_2
    global BDBRP_avg_3     
    global BDBR_avg_3        
    global Delta_YPSNR_avg_3
    global Delta_UPSNR_avg_3
    global Delta_VPSNR_avg_3
    global Delta_time_avg_3

    ## 1.计算BD-rate
    BDBR_P = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                        testYPSNR_dict[filename], testBit_dict[filename], True)
    BDBR_P = float('%.1f' %(BDBR_P * 100))

    #print index_num, filename, '-'*(50-len(filename)+10), str(float('%.1f'  %((BDBR_P)))) + '%'

    BDBR = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                        testYPSNR_dict[filename], testBit_dict[filename], False)
    BDBR = float('%.1f' %(BDBR * 100))
    BDBRP_avg += BDBR_P
    BDBR_avg  += BDBR
    if (len(sys.argv) == 4):
        print index_num, filename, '-'*(50-len(filename)+10), str(float('%.1f'  %((BDBR_P)))) + '%'

    if (len(sys.argv) > 4):
        BDBR_P_2 = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                        testYPSNR_dict2[filename], testBit_dict2[filename], True)
        BDBR_P_2 = float('%.1f' %(BDBR_P_2 * 100))

        BDBR_2 = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                        testYPSNR_dict2[filename], testBit_dict2[filename], False)
        BDBR_2 = float('%.1f' %(BDBR_2 * 100))
        BDBRP_avg_2 += BDBR_P_2
        BDBR_avg_2  += BDBR_2
        if(len(sys.argv) == 5):
            print index_num, filename, '-'*(50-len(filename)+10), str(float('%.1f'  %((BDBR_P)))) + '%', \
                    str(float('%.1f'  %((BDBR_P_2)))) + '%'
    if (len(sys.argv) > 5):
        BDBR_P_3 = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                    testYPSNR_dict3[filename], testBit_dict3[filename], True)
        BDBR_P_3 = float('%.1f' %(BDBR_P_3 * 100))

        BDBR_3 = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                    testYPSNR_dict3[filename], testBit_dict3[filename], False)
        BDBR_3 = float('%.1f' %(BDBR_3 * 100))
        BDBRP_avg_3 += BDBR_P_3
        BDBR_avg_3  += BDBR_3
        if (len(sys.argv) == 6):
            print index_num, filename, '-'*(50-len(filename)+10), str(float('%.1f'  %((BDBR_P)))) + '%', \
                str(float('%.1f'  %((BDBR_P_2)))) + '%', str(float('%.1f'  %((BDBR_P_3)))) + '%'  
        
    ## 2.计算Delta_YPSNR, Delta_UPSNR, Delta_VPSNR
    Delta_YPSNR_list = testYPSNR_dict[filename] - origYPSNR_dict[filename]
    for i in Delta_YPSNR_list:
        Delta_YPSNR = Delta_YPSNR + i
    Delta_YPSNR = float('%.3f' %(Delta_YPSNR / len(Delta_YPSNR_list)))
    
    Delta_UPSNR_list = testUPSNR_dict[filename] - origUPSNR_dict[filename]
    for i in Delta_UPSNR_list:
        Delta_UPSNR = Delta_UPSNR + i
    Delta_UPSNR = float('%.3f' %(Delta_UPSNR / len(Delta_UPSNR_list)))

    Delta_VPSNR_list = testVPSNR_dict[filename] - origVPSNR_dict[filename]
    for i in Delta_VPSNR_list:
        Delta_VPSNR = Delta_VPSNR + i
    Delta_VPSNR = float('%.3f' %(Delta_VPSNR / len(Delta_VPSNR_list)))

    Delta_YPSNR_avg += Delta_YPSNR
    Delta_UPSNR_avg += Delta_UPSNR
    Delta_VPSNR_avg += Delta_VPSNR

    if (len(sys.argv) > 4):
        Delta_YPSNR_list = testYPSNR_dict2[filename] - origYPSNR_dict[filename]
        for i in Delta_YPSNR_list:
            Delta_YPSNR_2 = Delta_YPSNR_2 + i
        Delta_YPSNR_2 = float('%.3f' %(Delta_YPSNR_2 / len(Delta_YPSNR_list)))
        
        Delta_UPSNR_list = testUPSNR_dict2[filename] - origUPSNR_dict[filename]
        for i in Delta_UPSNR_list:
            Delta_UPSNR_2 = Delta_UPSNR_2 + i
        Delta_UPSNR_2 = float('%.3f' %(Delta_UPSNR_2 / len(Delta_UPSNR_list)))

        Delta_VPSNR_list = testVPSNR_dict2[filename] - origVPSNR_dict[filename]
        for i in Delta_VPSNR_list:
            Delta_VPSNR_2 = Delta_VPSNR_2 + i
        Delta_VPSNR_2 = float('%.3f' %(Delta_VPSNR_2 / len(Delta_VPSNR_list)))

        Delta_YPSNR_avg_2 += Delta_YPSNR_2
        Delta_UPSNR_avg_2 += Delta_UPSNR_2
        Delta_VPSNR_avg_2 += Delta_VPSNR_2
    if (len(sys.argv) > 5):
        Delta_YPSNR_list = testYPSNR_dict3[filename] - origYPSNR_dict[filename]
        for i in Delta_YPSNR_list:
            Delta_YPSNR_3 = Delta_YPSNR_3 + i
        Delta_YPSNR_3 = float('%.3f' %(Delta_YPSNR_3 / len(Delta_YPSNR_list)))
        
        Delta_UPSNR_list = testUPSNR_dict3[filename] - origUPSNR_dict[filename]
        for i in Delta_UPSNR_list:
            Delta_UPSNR_3 = Delta_UPSNR_3 + i
        Delta_UPSNR_3 = float('%.3f' %(Delta_UPSNR_3 / len(Delta_UPSNR_list)))

        Delta_VPSNR_list = testVPSNR_dict3[filename] - origVPSNR_dict[filename]
        for i in Delta_VPSNR_list:
            Delta_VPSNR_3 = Delta_VPSNR_3 + i
        Delta_VPSNR_3 = float('%.3f' %(Delta_VPSNR_3 / len(Delta_VPSNR_list)))

        Delta_YPSNR_avg_3 += Delta_YPSNR_3
        Delta_UPSNR_avg_3 += Delta_UPSNR_3
        Delta_VPSNR_avg_3 += Delta_VPSNR_3
    
    ## 3.计算Delta_time
    Delta_time_list = ((testTime_dict[filename] - origTime_dict[filename])/ \
                            origTime_dict[filename]) *100
    for i in Delta_time_list:
        Delta_time = Delta_time + i
    Delta_time = float(Delta_time / len(Delta_time_list))
    Delta_time = float('%.3f' %(Delta_time))
    Delta_time_avg += Delta_time

    if (len(sys.argv) > 4):
        Delta_time_list = ((testTime_dict2[filename] - origTime_dict[filename])/ \
                            origTime_dict[filename]) *100
        for i in Delta_time_list:
            Delta_time_2 = Delta_time_2 + i
        Delta_time_2 = float(Delta_time_2 / len(Delta_time_list))
        Delta_time_2 = float('%.3f' %(Delta_time_2))
        Delta_time_avg_2 += Delta_time_2           
    if (len(sys.argv) > 5):
        Delta_time_list = ((testTime_dict3[filename] - origTime_dict[filename])/ \
                            origTime_dict[filename]) *100
        for i in Delta_time_list:
            Delta_time_3 = Delta_time_3 + i
        Delta_time_3 = float(Delta_time_3 / len(Delta_time_list))
        Delta_time_3 = float('%.3f' %(Delta_time_3))
        Delta_time_avg_3 += Delta_time_3   

    # 两路对比
    oneline = filename +' ' + str(BDBR_P) + ' ' + str(BDBR) + ' ' + str(Delta_YPSNR) + ' ' \
            + str(Delta_UPSNR) + ' ' + str(Delta_VPSNR) + ' ' + str(Delta_time) + '\n'
    if (len(sys.argv) > 4):  # 三路对比
        oneline = filename +' ' + str(BDBR_P) + ' ' + str(BDBR) + ' ' + str(Delta_YPSNR) + ' ' \
                + str(Delta_UPSNR) + ' ' + str(Delta_VPSNR) + ' ' + str(Delta_time) + ' '      \
                + str(BDBR_P_2) + ' ' + str(BDBR_2) + ' ' + str(Delta_YPSNR_2) + ' '           \
                + str(Delta_UPSNR_2) + ' ' + str(Delta_VPSNR_2) + ' ' + str(Delta_time_2) + '\n'    
    if (len(sys.argv) > 5): # 四路对比
        oneline = filename +' ' + str(BDBR_P) + ' ' + str(BDBR) + ' ' + str(Delta_YPSNR) + ' ' \
                + str(Delta_UPSNR) + ' ' + str(Delta_VPSNR) + ' ' + str(Delta_time) + ' '      \
                + str(BDBR_P_2) + ' ' + str(BDBR_2) + ' ' + str(Delta_YPSNR_2) + ' '           \
                + str(Delta_UPSNR_2) + ' ' + str(Delta_VPSNR_2) + ' ' + str(Delta_time_2) + ' '\
                + str(BDBR_P_3) + ' ' + str(BDBR_3) + ' ' + str(Delta_YPSNR_3) + ' '           \
                + str(Delta_UPSNR_3) + ' ' + str(Delta_VPSNR_3) + ' ' + str(Delta_time_3) + '\n'
    return oneline

####################################main 函数入口####################################################
if __name__ == '__main__':
    if(len(sys.argv) < 4):
        print('Usage: auto_data_analysis.py ' + '<anchor_vs._ref_result outDir>' + '\n')
        print("For example: auto_data_analysis.py anchor_vs._ref_result ./out ")
        print('Notice: <> is necessary, [] is optional')
        exit()
    
    ## 1.命令行参数解析    
    anchor = sys.argv[1]
    outDir = sys.argv[2]
    refer1 = sys.argv[3]
    
    if (len(sys.argv) > 4):
        refer2 = sys.argv[4]
        testBit_dict2   = collections.OrderedDict()
        testYPSNR_dict2 = collections.OrderedDict()
        testUPSNR_dict2 = collections.OrderedDict()
        testVPSNR_dict2 = collections.OrderedDict()
        testTime_dict2  = collections.OrderedDict()
    if (len(sys.argv) > 5):
        refer3 = sys.argv[5]
        testBit_dict3   = collections.OrderedDict()
        testYPSNR_dict3 = collections.OrderedDict()
        testUPSNR_dict3 = collections.OrderedDict()
        testVPSNR_dict3 = collections.OrderedDict()
        testTime_dict3  = collections.OrderedDict()
    
    ## 2.创建输出目录
    make_all_dir(outDir)

    ## 3.保存数据分析结果的字典  
    origBit_dict   = collections.OrderedDict()  ## key: seq_name value: bitrate
    origYPSNR_dict = collections.OrderedDict()
    origUPSNR_dict = collections.OrderedDict()
    origVPSNR_dict = collections.OrderedDict()
    testBit_dict   = collections.OrderedDict()
    testYPSNR_dict = collections.OrderedDict()
    testUPSNR_dict = collections.OrderedDict()
    testVPSNR_dict = collections.OrderedDict()
    origTime_dict  = collections.OrderedDict()
    testTime_dict  = collections.OrderedDict()
    seqName_dict   = collections.OrderedDict()  ## key: index_num  value: seq_name

    ## 4.读取anchor数据并提取bitrate,PSNR和time信息
    seq_num = shuffle_info(anchor, 1)

    ## 5.读取refer数据并提取bitrate,PSNR和time信息
    anchor_codec = anchor.split('__result')[1].split('_')[1]
    refer1_codec = anchor.split('__result')[1].split('_')[3]
    print anchor_codec, refer1_codec

    ## 6.创建数据分析结果csv文件
    outExcelData = outDir+delimiter+'__result_'+anchor_codec+ '_vs._'+refer1_codec+'_BDBR.csv'
    create_excel(outExcelData)

    pFile = open(outExcelData, 'w') #创建汇总文件，性能数据
    pFile.write(codecs.BOM_UTF8)
    csv_writer=csv.writer(pFile, dialect='excel')

    #写入输出结果Excel的标题，默认两路对比
    totaltitle=['video sequence', 'BD-rate(piecewise_cubic)(%)', 'BD-rate(cubic)(%)', 'Delta_Y-PSNR(dB)', 'Delta_U-PSNR(dB)', 'Delta_V-PSNR(dB)', 'Delta_time(%)']
    csv_writer.writerow(totaltitle)
    pFile.close()

    BDBRP_avg       = 0.0
    BDBR_avg        = 0.0
    Delta_YPSNR_avg = 0.0
    Delta_UPSNR_avg = 0.0
    Delta_VPSNR_avg = 0.0
    Delta_time_avg  = 0.0 

    ## 7. 计算BD-rate(piecewise cubic)和BD-rate(cubic)以及Delta_PSNR和Delta_time
    for index_num in range(1, seq_num + 1):        
        filename = seqName_dict[index_num]
        print filename

        ## 7.1 计算BD-rate,DeltaPSNR和DeltaTime
        oneline = ComputeBDBR_PSNR_Time(index_num)
        
        ## 7.2 保存数据
        pFile = open(outExcelData, 'a+')
        pFile.write(codecs.BOM_UTF8)
        csv_writer=csv.writer(pFile, dialect='excel')
        csv_writer.writerow(oneline.split())
        pFile.close()

    ## 8. 输出平均值信息
    BDBRP_avg       = float('%.1f' %(BDBRP_avg   / seq_num))
    BDBR_avg        = float('%.1f' %(BDBR_avg    / seq_num))
    Delta_YPSNR_avg = float('%.3f' %(Delta_YPSNR_avg / seq_num))
    Delta_UPSNR_avg = float('%.3f' %(Delta_UPSNR_avg / seq_num))
    Delta_VPSNR_avg = float('%.3f' %(Delta_VPSNR_avg / seq_num))
    Delta_time_avg  = float('%.3f' %(Delta_time_avg  / seq_num))
 
    pFile = open(outExcelData, 'a+') #创建汇总文件，追加性能数据
    pFile.write(codecs.BOM_UTF8)
    csv_writer=csv.writer(pFile, dialect='excel')

    average_data = 'Average:\n' + str(BDBRP_avg) + ' ' + str(BDBR_avg) + ' ' + str(Delta_YPSNR_avg)  \
                 + ' ' + str(Delta_UPSNR_avg) + ' ' + str(Delta_VPSNR_avg) + ' ' + str(Delta_time_avg)
    average_data_show = 'Summary:\n' + anchor_codec + ' vs. ' + refer1_codec + ': ' + str(BDBRP_avg) + ' ' \
                + str(BDBR_avg) + ' ' + str(Delta_YPSNR_avg) + ' ' + str(Delta_UPSNR_avg) + ' '            \
                + str(Delta_VPSNR_avg) + ' ' + str(Delta_time_avg)
                                             
    print '\n', average_data_show
    csv_writer.writerow(average_data.split())
    pFile.close()

    ## 9. 将csv文件转换成excel文件,此处为了在一个表格里面绘制率失真曲线图
    analysis_file = outDir+delimiter+'analysis_result_'+anchor_codec+'_vs._'+refer1_codec+'.xlsx'

    writer    = pd.ExcelWriter(analysis_file)

    csv_file1 = pd.read_csv(outExcelData, encoding='utf-8')
    csv_file1.to_excel(writer, sheet_name='result')

    csv_file2 = pd.read_csv(anchor, encoding='utf-8')
    #csv_file2.to_excel(writer, sheet_name='anchor_data')

 
    writer.save()

    wb = load_workbook(analysis_file)
    #print wb.get_sheet_names()
    sheet_result = wb.get_sheet_by_name('result')  # 获得当前正在显示的sheet
    #sheet_anchor = wb.get_sheet_by_name('anchor_data')  # 获得当前正在显示的sheet
    #sheet_refer  = wb.get_sheet_by_name('refer1_data')  # 获得当前正在显示的sheet    

    #excelCurrRow = 2
    #dataVerStep  = 4
    ## 10. 绘制率失真曲线图
    # for index_num in range(1, seq_num + 1):
    #     filename = seqName_dict[index_num] 
    #     line = openpyxl.chart.ScatterChart()
    #     line.title = filename                   #图表标题
    #     line.x_axis.title = 'Bitrate (kbps)'    #y轴标题
    #     line.y_axis.title = 'Y-PSNR (dB)'       #x轴标题

    #     line.y_axis.scaling.min = int(min(origYPSNR_dict[filename][3], testYPSNR_dict[filename][3])) - 2  # y轴的最小值
        
    #     oriXdata = Reference(sheet_anchor, min_col=4, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
    #     oriYdata = Reference(sheet_anchor, min_col=5, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
    #     series = Series(oriYdata, oriXdata, title=anchor_codec)
    #     series.marker.symbol = 'circle'
    #     series.smooth = True
    #     line.series.append(series)

    #     testXdata = Reference(sheet_refer, min_col=4, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
    #     testYdata = Reference(sheet_refer, min_col=5, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
    #     series = Series(testYdata, testXdata, title=refer1_codec)
    #     series.marker.symbol = 'circle'
    #     series.smooth = True
    #     line.series.append(series)

    #     # 两路对比
    #     if (index_num-1) % 2 == 0:
    #         chartColumn = 'J'
    #     else:
    #         chartColumn = 'T'
        
    #     sheet_result.add_chart(line, chartColumn + str(excelCurrRow + (index_num-1) * (dataVerStep+4)+2))       
    # 保存分析结果文件
    wb.save(analysis_file)
    
    ret = 0
    if(ret != -1):
        print("---------Process finished!---------")
        os._exit(0)