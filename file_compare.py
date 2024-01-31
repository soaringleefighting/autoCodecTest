#_*_ coding=utf-8_*_  #脚本中有中文注释必须包含这一句

#######################################################################################
##脚本功能： 本脚本用于比较文件内容
##脚本用法： python file_compare.py file_dir1 file_dir2 compare_output file_type
##参数说明：    file_dir1     :	  文件夹1
##             file_dir2             :   文件夹2
##             compare_output      :   比较结果
##             file_type        :   文件类型(若为所有文件则输入"all",不用加点)
##             注意：请将文件顺序保持一致，该程序仅按照顺序依次进行
## Created by gaojinzheng at 2023.8.10
import filecmp
import os
import sys
import openpyxl
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf-8')

def compare(file1,file2):
    result = filecmp.cmp(file1,file2)
    return result

def compare_all(dir1,dir2):
    result = filecmp.dircmp(dir1,dir2)
    return result

def get_file_from_dir(dir,file_type):
    file_list = []
    if(file_type == 'all') :
        for curDir, dirs, files in os.walk(dir):
            for file in files:
                print(os.path.join(curDir, file))
                file_list.append(os.path.join(curDir, file))
    else:
        for curDir, dirs, files in os.walk(dir):
            for file in files:
                if file.endswith('.'+file_type):
                    print(os.path.join(curDir, file))
                    file_list.append(os.path.join(curDir, file))
                    
    return file_list

if __name__ == "__main__":
    if(len(sys.argv)<=4):
        print('输入参数数目不足,请依次输入要比对的文件夹1 文件夹2 储存地址 文件类型(若需要全部比较,则输入all)')
        exit()
    else:
        file_dir1 = sys.argv[1]
        file_dir2 = sys.argv[2]
        file_save = sys.argv[3]
        file_type = sys.argv[4]
        print('文件夹1路径为{}\n文件夹2路径为{}\n文件类型为{}\n'.format(file_dir1,file_dir2,file_type))
    print('------正在读取文件夹{}------\n'.format(file_dir1))
    file1_list = get_file_from_dir(file_dir1,file_type)
    if(len(file1_list) == 0):
        print('读取错误！')
        exit()
    else:
        print('读取完成,共有{}个文件!'.format(len(file1_list)))
    print('------正在读取文件夹{}------\n'.format(file_dir2))
    file2_list = get_file_from_dir(file_dir2,file_type)
    if(len(file2_list) == 0):
        print('读取错误！')
        exit()
    else:
        print('------读取完成,共有{}个文件!------'.format(len(file2_list)))
    
    if (file_type == 'all'):
        result_list = compare_all(file_dir1,file_dir2)
        result_list.report()
        print('------请注意：该模式直接打印结果,无法保存,如果需要保存,请使用hashlib库!------')
    else:
        result_list = []
        file_list = zip(file1_list,file2_list)
        for (file1,file2) in file_list:
            compare_result = compare(file1,file2)
            print('正在比较文件{}(文件夹1)和{}(文件夹2)\n比较结果:{}'.format(file1,file2,compare_result))
            result_list.append(compare_result)

        print('------正在保存结果------')
        wb = Workbook()
        sheet = wb.create_sheet(index=0,title='compare')

        sheet['A1'].value = '文件夹1'
        sheet['b1'].value = '文件夹2'
        sheet['c1'].value = '比较结果'

        sheet_list = zip(range(2,len(file1_list)+2),file1_list,file2_list,result_list)
        for (row,file1,file2,result) in sheet_list:
            sheet.cell(row=row,column=1).value = file1
            sheet.cell(row=row,column=2).value = file2
            sheet.cell(row=row,column=3).value = result
        
        wb.save(file_save + '/compare_result.xlsx')
        print('结果保存完成,请查看文件compare_result.xlsx')
        num_true = 0
        num_false = 0
        for result in result_list:
            if(result):
                num_true+=1
            else:
                num_false+=1
        print("结果统计：总数量——{}，一致数量——{}，不一致数量：——{}".format(len(result_list),num_true,num_false))