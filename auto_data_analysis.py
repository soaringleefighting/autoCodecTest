#_*_ coding=utf-8_*_  #脚本中有中文注释必须包含这一句

#######################################################################################
##脚本功能： 本脚本用于计算BD-rate, Delta_PSNR, SSIM-BDBR, Delta_time, 并绘制率失真曲线图。
##脚本用法： python auto_data_analysis.py anchor_result ./out SSIM refer1_result [refer2_result refer3_result]
##参数说明：    anchor_result     :	  原始测试数据
##             refer_result      :   对比测试数据
##             ./out             :   输出结果目录
##             SSIM              :   1:分析SSIM 0: 不分析SSIM 
##
## Created by lipeng at July 10 2020
## Version 2.3
## Modified:
## 2020.7.10 create tag v1.0 first version
## 2020.7.16 create tag v2.0 support BDBR collect
## 2021.7.30 create tag v2.1 support calculate BDBR and plot RD curve
## 2021.7.31 create tag v2.2 支持三路或四路编码器对比计算BD-rate和绘制率失真曲线图(绘制在一张图上)
## 2024.1.31 create tag V2.4 支持SSIM-BDBR
#######################################################################################
import os
import re
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import glob
import filecmp
import shutil
import subprocess
import subprocess as sub
import csv
import codecs
from   collections import OrderedDict
import collections
import xlrd
import xlwt
from   xlutils.copy import copy
#import scipy 							#特殊三对角矩阵求解
import numpy as np 						#用于数学计算(曲线拟合、积分等)
import matplotlib.pyplot as plt         #绘制率失真曲线
import pandas as pd                     #转换excel
import openpyxl                         #Excel文件
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl import load_workbook

space     = ' '
delimiter = '/'

reload(sys)
sys.setdefaultencoding('utf-8')

#比较两个文件是否相同, 相同则返回True, 不同返回False
def	yuv_cmp(file1,file2):
	isNul1 = os.path.getsize(file1)
	isNul2 = os.path.getsize(file2)
	if((not isNul1) or (not isNul2)):
		return False
	if(isNul1 == isNul2):
		return True

#提取文件的名字
def get_file_name(fullfilename):
	tmp = fullfilename.strip()
	name = os.path.split(tmp)[-1]   #提取文件名，不包含路径
	return os.path.splitext(name)[0] #提取文件名，不包含后缀

#创建文件目录
def make_all_dir(path):	
	path = path.strip() #去除首位空格
	path=path.rstrip("\\")  #去除尾部\符号

	isExist = os.path.exists(path) #判断路径是否存在
        if not isExist:        #如果不存在则创建目录
                os.makedirs(path)
                print '[info]: ' + path+' Create success!'
                return True
        else:   #如果目录存在则不创建，并提示目录已经存在
                print '[info]: ' + path+' aleady exist!'
                return False

#获取数据文件
def	get_raw_data(rawdir):
	isfile = 0
	if os.path.isdir(rawdir):
		allfiles = os.listdir(rawdir)
		files = [rawdir+"/"+f for f in allfiles if re.search('txt$',f) or  re.search('log$',f)]
	elif os.path.isfile(rawdir):
		isfile = 1
		files = [rawdir]
	else:
		files = []
		print("ERROR: " + sys.argv[1] + "  is not a dir or file!")
	files.sort(key=str.lower)
	return [files,isfile]

#create new excel file
def create_excel(excel_name):
    file = open(excel_name,'wb')
    writer=csv.writer(file)
    file.close()


#collect data from format text to BDBR excel for ref data
def collect_data_to_BDBRexcel(exceldata, datawt, inputfile, outexcel):
    pFile = open(inputfile, 'a+')
    lines = pFile.readlines()
    data = collections.OrderedDict()  ##使用有序字典
    splitValue = lines[0].strip().split(' ')
    data[splitValue[0]] = [splitValue[0], splitValue[1], splitValue[2], splitValue[3]]
    sequence_name_plus_qp = splitValue[0]
    sequence_name = splitValue[0].split('_qp')[0]
    sequence_qp = splitValue[0].split('_qp')[1]

    table = exceldata.sheet_by_name('AI-Main')
    table_wt = datawt.get_sheet('AI-Main')

    ##遍历excel中每一行，存在匹配的字符串则写入对应的bitrate,Y-PSNR和EncT
    nrows = table.nrows
    for i in range(nrows):
        if type(table.col_values(2)[i]) == float:  ##将float类型转换成int类型
            qp = int(table.col_values(2)[i])
        if str(table.col_values(1)[i]) == sequence_name and str(qp) == sequence_qp:
            table_wt.write(i, 11, splitValue[1]) #write bitrate
            table_wt.write(i, 12, splitValue[2]) #write Y-PSNR
            table_wt.write(i, 15, splitValue[3]) #write EncT(s)
    datawt.save(outexcel)
    return 0


# collect data from format text to BDBR excel for anchor data
def collect_data_to_BDBRexcel_vs(exceldata, datawt, inputfile, outexcel):
    pFile = open(inputfile, 'a+')
    lines = pFile.readlines()
    data = collections.OrderedDict()  ##使用有序字典
    splitValue = lines[0].strip().split(' ')
    data[splitValue[0]] = [splitValue[0], splitValue[1], splitValue[2], splitValue[3]]
    sequence_name_plus_qp = splitValue[0]
    sequence_name = splitValue[0].split('_qp')[0]
    sequence_qp = splitValue[0].split('_qp')[1]

    table = exceldata.sheet_by_name('AI-Main')
    table_wt = datawt.get_sheet('AI-Main')

    ##遍历excel中每一行，存在匹配的字符串则写入对应的bitrate,Y-PSNR和EncT
    nrows = table.nrows
    for i in range(nrows):
        if type(table.col_values(2)[i]) == float:  ##将float类型转换成int类型
            qp = int(table.col_values(2)[i])
        if str(table.col_values(1)[i]) == sequence_name and str(qp) == sequence_qp:
            table_wt.write(i, 3, splitValue[1]) #write bitrate
            table_wt.write(i, 4, splitValue[2]) #write Y-PSNR
            table_wt.write(i, 7, splitValue[3]) #write EncT(s)
            ## 计算编码节省时间
    datawt.save(outexcel)
    return 0

## 从特定格式csv文件中提取码率,PSNR,SSIM和time信息
def shuffle_info(anchor, isAnchor=1, refer_idx=0):
    anchordata    = csv.reader(open(anchor), quotechar="'") ## 'r'
    count_num = -1
    index_num =  0

    origBit_arr    = np.zeros(4)
    testBit_arr    = np.zeros(4)
    origYPSNR_arr  = np.zeros(4)
    testYPSNR_arr  = np.zeros(4)
    origUPSNR_arr  = np.zeros(4)
    testUPSNR_arr  = np.zeros(4)
    origVPSNR_arr  = np.zeros(4)
    testVPSNR_arr  = np.zeros(4)
    origYSSIM_arr  = np.zeros(4)
    testYSSIM_arr  = np.zeros(4)
    origTime_arr   = np.zeros(4)
    testTime_arr   = np.zeros(4)

    for anchor_line in anchordata:
        count_num = count_num + 1
        if count_num == 0:
            continue
        seq_name = '_'.join([anchor_line[0].split('_')[0], anchor_line[0].split('_')[1]])
        bitrate = anchor_line[2]
        Y_PSNR  = anchor_line[3]
        U_PSNR  = anchor_line[4]
        V_PSNR  = anchor_line[5]
        Y_SSIM  = anchor_line[6]
        time    = anchor_line[-1]  # time
        if isAnchor == 1:
            origBit_arr[index_num]    = bitrate
            origYPSNR_arr[index_num]  = Y_PSNR
            origUPSNR_arr[index_num]  = U_PSNR
            origVPSNR_arr[index_num]  = V_PSNR
            origYSSIM_arr[index_num]  = Y_SSIM
            origTime_arr[index_num]   = time
        else:
            testBit_arr[index_num]    = bitrate
            testYPSNR_arr[index_num]  = Y_PSNR
            testUPSNR_arr[index_num]  = U_PSNR
            testVPSNR_arr[index_num]  = V_PSNR
            testYSSIM_arr[index_num]  = Y_SSIM
            testTime_arr[index_num]   = time    
        index_num=index_num+1
        if index_num == 4:
            index_num = 0
            if isAnchor == 1:
                origBit_dict[seq_name]  = origBit_arr.copy()  #深拷贝
                origYPSNR_dict[seq_name] = origYPSNR_arr.copy()
                origUPSNR_dict[seq_name] = origUPSNR_arr.copy()
                origVPSNR_dict[seq_name] = origVPSNR_arr.copy()
                origYSSIM_dict[seq_name] = origYSSIM_arr.copy()
                origTime_dict[seq_name] = origTime_arr.copy()
                seqName_dict[count_num/4] = seq_name
            else:
                if refer_idx == 0:
                    testBit_dict  [seq_name] = testBit_arr.copy()
                    testYPSNR_dict[seq_name] = testYPSNR_arr.copy()
                    testUPSNR_dict[seq_name] = testUPSNR_arr.copy()
                    testVPSNR_dict[seq_name] = testVPSNR_arr.copy()
                    testYSSIM_dict[seq_name] = testYSSIM_arr.copy()
                    testTime_dict [seq_name] = testTime_arr.copy()
                elif refer_idx == 1:
                    testBit_dict2  [seq_name] = testBit_arr.copy()
                    testYPSNR_dict2[seq_name] = testYPSNR_arr.copy()
                    testUPSNR_dict2[seq_name] = testUPSNR_arr.copy()
                    testVPSNR_dict2[seq_name] = testVPSNR_arr.copy()
                    testYSSIM_dict2[seq_name] = testYSSIM_arr.copy()
                    testTime_dict2 [seq_name] = testTime_arr.copy() 
                elif refer_idx == 2:
                    testBit_dict3  [seq_name] = testBit_arr.copy()
                    testYPSNR_dict3[seq_name] = testYPSNR_arr.copy()
                    testUPSNR_dict3[seq_name] = testUPSNR_arr.copy()
                    testVPSNR_dict3[seq_name] = testVPSNR_arr.copy()
                    testYSSIM_dict3[seq_name] = testYSSIM_arr.copy()
                    testTime_dict3 [seq_name] = testTime_arr.copy() 
    return (count_num/4)

def csv_to_xlsx(csvfile, excelfile):
    with open(csvfile, 'r') as f:
        read = csv.reader(codecs.EncodedFile(f, 'utf8', 'utf_8_sig'), delimiter=" ")
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('result')  # 创建一个sheet表格
        l = 0
        for line in read:
            #print(line)
            r = 0
            for i in line:
                #print(i)
                sheet.write(l, r, unicode(i, 'utf-8'))  # 一个一个将单元格数据写入
                r = r + 1
            l = l + 1
        workbook.save(excelfile)  # 保存Excel

def pchip_end(h1, h2, del1, del2):
    d = ((2*h1 + h2)*del1 - h1*del2) / (h1 + h2)
    if np.sign(d) != np.sign(del1):
        d = 0
    elif np.sign(del1) != np.sign(del2) and np.abs(d) > np.abs(3*del1):
        d = 3 * del1
    return d

def pchip_slopes(h, delta):
    d = np.zeros(len(h) + 1)
    k = np.argwhere(np.sign(delta[:-1]) * np.sign(delta[1:]) > 0).reshape(-1) + 1
    w1 = 2*h[k] + h[k-1]
    w2 = h[k] + 2*h[k-1]
    d[k] = (w1 + w2) / (w1 / delta[k-1] + w2 / delta[k])
    d[0] = pchip_end(h[0], h[1], delta[0], delta[1])
    d[-1] = pchip_end(h[-1], h[-2], delta[-1], delta[-2])
    return d

def spline_slopes(h, delta):
    a, r = np.zeros([3, len(h)+1]), np.zeros(len(h)+1)
    a[0, 1], a[0, 2:] = h[0] + h[1], h[:-1]
    a[1, 0], a[1, 1:-1], a[1, -1] = h[1], 2*(h[:-1] + h[1:]), h[-2]
    a[2, :-2], a[2, -2] = h[1:], h[-1] + h[-2]

    r[0] = ((h[0] + 2*a[0, 1])*h[1]*delta[0] + h[0]**2*delta[1]) / a[0, 1]
    r[1:-1] = 3*(h[1:] * delta[:-1] + h[:-1] * delta[1:])
    r[-1] = (h[-1]**2*delta[-2] + (2*a[2, -2] + h[-1])*h[-2]*delta[-1]) / a[2, -2]

    d = scipy.linalg.solve_banded((1, 1), a, r)
    return d

class PCHIP:
    def __init__(self, x, y, use_spline=False):
        assert len(np.unique(x)) == len(x)
        #将数据按x坐标从小到大排序
        order = np.argsort(x)
        self.xi, self.yi = x[order], y[order]

        #求输入x、y的间隔，按间隔计算delta
        h = np.diff(self.xi)
        delta = np.diff(self.yi) / h

        self.d = spline_slopes(h, delta) if use_spline else pchip_slopes(h, delta)
        self.c = (3*delta - 2*self.d[:-1] - self.d[1:]) / h
        self.b = (self.d[:-1] - 2*delta + self.d[1:]) / h**2

        """
        The piecewise function is like p(x) = y_k + s*d_k + s*s*c_k + s*s*s*b_k
        where s = x - xi_k, k is the interval includeing x.
        So the original function of p(x) is P(x) = s*y_k + 1/2*s*s*d_k + 1/3*s*s*s*c_k + 1/4*s*s*s*s*b_k + C.
        """
        self.interval_int_coeff = []
        self.interval_int = np.zeros(len(x)-1)
        for i in range(len(x)-1):
            self.interval_int_coeff.append(np.polyint([self.b[i], self.c[i], self.d[i], self.yi[i]]))
            self.interval_int[i] = np.polyval(self.interval_int_coeff[-1], h[i]) - np.polyval(self.interval_int_coeff[-1], 0)

    def _integral(self, lower, upper):
        assert lower <= upper
        if lower < np.min(self.xi):
            lower = np.min(self.xi)
            print('Warning: The lower bound is less than the interval and clipped!')
        elif lower > np.max(self.xi):
            print('Warning: The lower bound is greater than the interval!')
            return 0
        if upper > np.max(self.xi):
            upper = np.max(self.xi)
            print('Warning: The upper bound is greater than the interval and clipped!')
        elif upper < np.min(self.xi):
            print('Warning: The lower bound is less than the interval!')
            return 0
        left = np.arange(len(self.xi))[self.xi - lower > -1e-6][0]
        right = np.arange(len(self.xi))[self.xi - upper < 1e-6][-1]

        inte = np.sum(self.interval_int[left:right])
        if self.xi[left] - lower > 1e-6:
            inte += (np.polyval(self.interval_int_coeff[left-1], self.xi[left]-self.xi[left-1]) - np.polyval(self.interval_int_coeff[left-1], lower-self.xi[left-1]))
        if self.xi[right] - upper < -1e-6:
            inte += (np.polyval(self.interval_int_coeff[right], upper-self.xi[right]) - np.polyval(self.interval_int_coeff[right], 0))
        return inte

    def integral(self, lower, upper):
        if lower > upper:
            return -self._integral(upper, lower)
        else:
            return self._integral(lower, upper)

def computeBDRate(testNum, basePSNR, baseBitrate, testPSNR, testBitrate, piecewise=True):
    #码率取对数
    baseLogBitrate = np.log10(baseBitrate)
    testLogBitrate = np.log10(testBitrate)
    #确定共同范围
    minPSNR = np.max((np.min(basePSNR), np.min(testPSNR)))
    maxPSNR = np.min((np.max(basePSNR), np.max(testPSNR)))
    if piecewise == True:
        baseIntegral = PCHIP(basePSNR, baseLogBitrate, use_spline=False).integral(minPSNR, maxPSNR)
        testIntegral = PCHIP(testPSNR, testLogBitrate, use_spline=False).integral(minPSNR, maxPSNR)
    else:
        #拟合曲线
        baseFitting = np.polyfit(basePSNR, baseLogBitrate, testNum-1)
        testFitting = np.polyfit(testPSNR, testLogBitrate, testNum-1)
        #不定积分
        baseIndefiniteInt = np.polyint(baseFitting)
        testIndefiniteInt = np.polyint(testFitting)
        #求积分
        baseIntegral = np.polyval(baseIndefiniteInt, maxPSNR) - np.polyval(baseIndefiniteInt, minPSNR)
        testIntegral = np.polyval(testIndefiniteInt, maxPSNR) - np.polyval(testIndefiniteInt, minPSNR)

    #平均差值
    meanEXPDiff = (testIntegral - baseIntegral) / (maxPSNR - minPSNR)
    meanDiff = (np.power(10, meanEXPDiff) - 1)

    return meanDiff


def ComputeBDBR_PSNR_SSIM_Time(index_num, SSIM):
    Delta_YPSNR        = 0.0
    Delta_UPSNR        = 0.0
    Delta_VPSNR        = 0.0
    Delta_time         = 0.0
    Delta_YPSNR_2      = 0.0
    Delta_UPSNR_2      = 0.0
    Delta_VPSNR_2      = 0.0
    Delta_time_2       = 0.0
    Delta_YPSNR_3      = 0.0
    Delta_UPSNR_3      = 0.0
    Delta_VPSNR_3      = 0.0
    Delta_time_3       = 0.0    
    
    global BDBRY_avg 
    global BDBRU_avg
    global BDBRV_avg       
    global YUVBDBR_avg
    global SSIMBDBR_avg 
    global Delta_YPSNR_avg 
    global Delta_UPSNR_avg 
    global Delta_VPSNR_avg
    global Delta_time_avg 
 
    global BDBRY_avg_2    
    global BDBRU_avg_2
    global BDBRV_avg_2     
    global YUVBDBR_avg_2
    global SSIMBDBR_avg_2 
    global Delta_YPSNR_avg_2
    global Delta_UPSNR_avg_2
    global Delta_VPSNR_avg_2 
    global Delta_time_avg_2
 
    global BDBRY_avg_3          
    global BDBRU_avg_3          
    global BDBRV_avg_3
    global YUVBDBR_avg_3
    global SSIMBDBR_avg_3 
    global Delta_YPSNR_avg_3
    global Delta_UPSNR_avg_3
    global Delta_VPSNR_avg_3
    global Delta_time_avg_3

    global SSIM_BDBRY
    global SSIM_BDBRY_2
    global SSIM_BDBRY_3
  
    ## 1.计算BD-rate(PSNR, SSIM)
    # BDBR piecewise cubic
    BDBR_Y = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                        testYPSNR_dict[filename], testBit_dict[filename], True)
    BDBR_U = computeBDRate(4, origUPSNR_dict[filename], origBit_dict[filename], \
                        testUPSNR_dict[filename], testBit_dict[filename], True)
    BDBR_V = computeBDRate(4, origVPSNR_dict[filename], origBit_dict[filename], \
                        testVPSNR_dict[filename], testBit_dict[filename], True)
    if (SSIM == 1):
        SSIM_BDBRY = computeBDRate(4, origYSSIM_dict[filename], origBit_dict[filename], \
                            testYSSIM_dict[filename], testBit_dict[filename], True)
    BDBR_Y = float('%.1f' %(BDBR_Y * 100))
    BDBR_U = float('%.1f' %(BDBR_U * 100))
    BDBR_V = float('%.1f' %(BDBR_V * 100))
    BDBRY_avg += BDBR_Y
    BDBRU_avg += BDBR_U
    BDBRV_avg += BDBR_V
    YUVBDBR_avg = float('%.3f' %((6 * BDBR_Y +  BDBR_U + BDBR_V) / 8))
    if (SSIM == 1):
        SSIM_BDBRY  = float('%.3f' %(SSIM_BDBRY * 100))
        SSIMBDBR_avg += SSIM_BDBRY

    if (len(sys.argv) > 5):
        BDBR_Y_2 = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                        testYPSNR_dict2[filename], testBit_dict2[filename], True)
        BDBR_U_2 = computeBDRate(4, origUPSNR_dict[filename], origBit_dict[filename], \
                        testUPSNR_dict2[filename], testBit_dict2[filename], True)
        BDBR_V_2 = computeBDRate(4, origVPSNR_dict[filename], origBit_dict[filename], \
                        testVPSNR_dict2[filename], testBit_dict2[filename], True)
        if (SSIM == 1):
            SSIM_BDBRY_2 = computeBDRate(4, origYSSIM_dict[filename], origBit_dict[filename], \
                            testYSSIM_dict2[filename], testBit_dict[filename], True)
        BDBR_Y_2 = float('%.1f' %(BDBR_Y_2 * 100))
        BDBR_U_2 = float('%.1f' %(BDBR_U_2 * 100))
        BDBR_V_2 = float('%.1f' %(BDBR_V_2 * 100))
        BDBRY_avg_2 += BDBR_Y_2
        BDBRU_avg_2 += BDBR_U_2
        BDBRV_avg_2 += BDBR_V_2
        YUVBDBR_avg_2 = float('%.3f' %((6 * BDBR_Y_2 +  BDBR_U_2 + BDBR_V_2) / 8))
        if (SSIM == 1):
            SSIM_BDBRY_2  = float('%.3f' %(SSIM_BDBRY_2 * 100))
            SSIMBDBR_avg_2 += SSIM_BDBRY_2
    if (len(sys.argv) > 6):
        BDBR_Y_3 = computeBDRate(4, origYPSNR_dict[filename], origBit_dict[filename], \
                    testYPSNR_dict3[filename], testBit_dict3[filename], True)
        BDBR_U_3 = computeBDRate(4, origUPSNR_dict[filename], origBit_dict[filename], \
                    testUPSNR_dict3[filename], testBit_dict3[filename], True)
        BDBR_V_3 = computeBDRate(4, origVPSNR_dict[filename], origBit_dict[filename], \
                    testVPSNR_dict3[filename], testBit_dict3[filename], True)
        if (SSIM == 1):
            SSIM_BDBRY_3 = computeBDRate(4, origYSSIM_dict[filename], origBit_dict[filename], \
                            testYSSIM_dict3[filename], testBit_dict[filename], True)
        BDBR_Y_3 = float('%.1f' %(BDBR_Y_3 * 100))
        BDBR_U_3 = float('%.1f' %(BDBR_U_3 * 100))
        BDBR_V_3 = float('%.1f' %(BDBR_V_3 * 100))
        BDBRY_avg_3 += BDBR_Y_3
        BDBRU_avg_3 += BDBR_U_3
        BDBRV_avg_3 += BDBR_U_3 
        YUVBDBR_avg_3 = float('%.3f' %((6 * BDBR_Y_3 +  BDBR_U_3 + BDBR_V_3) / 8)) 
        if (SSIM == 1):
            SSIM_BDBRY_3  = float('%.3f' %(SSIM_BDBRY_3 * 100))
            SSIMBDBR_avg_3 += SSIM_BDBRY_3
    ## 2.计算Delta_YPSNR, Delta_UPSNR, Delta_VPSNR
    Delta_YPSNR_list = testYPSNR_dict[filename] - origYPSNR_dict[filename]
    for i in Delta_YPSNR_list:
        Delta_YPSNR = Delta_YPSNR + i
    Delta_YPSNR = float('%.3f' %(Delta_YPSNR / len(Delta_YPSNR_list)))
    
    Delta_UPSNR_list = testUPSNR_dict[filename] - origUPSNR_dict[filename]
    for i in Delta_UPSNR_list:
        Delta_UPSNR = Delta_UPSNR + i
    Delta_UPSNR = float('%.3f' %(Delta_UPSNR / len(Delta_UPSNR_list)))

    Delta_VPSNR_list = testVPSNR_dict[filename] - origVPSNR_dict[filename]
    for i in Delta_VPSNR_list:
        Delta_VPSNR = Delta_VPSNR + i
    Delta_VPSNR = float('%.3f' %(Delta_VPSNR / len(Delta_VPSNR_list)))

    Delta_YPSNR_avg += Delta_YPSNR
    Delta_UPSNR_avg += Delta_UPSNR
    Delta_VPSNR_avg += Delta_VPSNR

    if (len(sys.argv) > 5):
        Delta_YPSNR_list = testYPSNR_dict2[filename] - origYPSNR_dict[filename]
        for i in Delta_YPSNR_list:
            Delta_YPSNR_2 = Delta_YPSNR_2 + i
        Delta_YPSNR_2 = float('%.3f' %(Delta_YPSNR_2 / len(Delta_YPSNR_list)))
        
        Delta_UPSNR_list = testUPSNR_dict2[filename] - origUPSNR_dict[filename]
        for i in Delta_UPSNR_list:
            Delta_UPSNR_2 = Delta_UPSNR_2 + i
        Delta_UPSNR_2 = float('%.3f' %(Delta_UPSNR_2 / len(Delta_UPSNR_list)))

        Delta_VPSNR_list = testVPSNR_dict2[filename] - origVPSNR_dict[filename]
        for i in Delta_VPSNR_list:
            Delta_VPSNR_2 = Delta_VPSNR_2 + i
        Delta_VPSNR_2 = float('%.3f' %(Delta_VPSNR_2 / len(Delta_VPSNR_list)))

        Delta_YPSNR_avg_2 += Delta_YPSNR_2
        Delta_UPSNR_avg_2 += Delta_UPSNR_2
        Delta_VPSNR_avg_2 += Delta_VPSNR_2
    if (len(sys.argv) > 6):
        Delta_YPSNR_list = testYPSNR_dict3[filename] - origYPSNR_dict[filename]
        for i in Delta_YPSNR_list:
            Delta_YPSNR_3 = Delta_YPSNR_3 + i
        Delta_YPSNR_3 = float('%.3f' %(Delta_YPSNR_3 / len(Delta_YPSNR_list)))
        
        Delta_UPSNR_list = testUPSNR_dict3[filename] - origUPSNR_dict[filename]
        for i in Delta_UPSNR_list:
            Delta_UPSNR_3 = Delta_UPSNR_3 + i
        Delta_UPSNR_3 = float('%.3f' %(Delta_UPSNR_3 / len(Delta_UPSNR_list)))

        Delta_VPSNR_list = testVPSNR_dict3[filename] - origVPSNR_dict[filename]
        for i in Delta_VPSNR_list:
            Delta_VPSNR_3 = Delta_VPSNR_3 + i
        Delta_VPSNR_3 = float('%.3f' %(Delta_VPSNR_3 / len(Delta_VPSNR_list)))

        Delta_YPSNR_avg_3 += Delta_YPSNR_3
        Delta_UPSNR_avg_3 += Delta_UPSNR_3
        Delta_VPSNR_avg_3 += Delta_VPSNR_3
    
    ## 3.计算Delta_time
    Delta_time_list = ((testTime_dict[filename] - origTime_dict[filename])/ \
                            origTime_dict[filename]) *100
    for i in Delta_time_list:
        Delta_time = Delta_time + i
    Delta_time = float(Delta_time / len(Delta_time_list))
    Delta_time = float('%.3f' %(Delta_time))
    Delta_time_avg += Delta_time

    if (len(sys.argv) > 5):
        Delta_time_list = ((testTime_dict2[filename] - origTime_dict[filename])/ \
                            origTime_dict[filename]) *100
        for i in Delta_time_list:
            Delta_time_2 = Delta_time_2 + i
        Delta_time_2 = float(Delta_time_2 / len(Delta_time_list))
        Delta_time_2 = float('%.3f' %(Delta_time_2))
        Delta_time_avg_2 += Delta_time_2           
    if (len(sys.argv) > 6):
        Delta_time_list = ((testTime_dict3[filename] - origTime_dict[filename])/ \
                            origTime_dict[filename]) *100
        for i in Delta_time_list:
            Delta_time_3 = Delta_time_3 + i
        Delta_time_3 = float(Delta_time_3 / len(Delta_time_list))
        Delta_time_3 = float('%.3f' %(Delta_time_3))
        Delta_time_avg_3 += Delta_time_3   

    # 打印到控制台
    if (SSIM == 1):
        print "{:2d}".format(index_num), filename, '-'*(50-len(filename)), "{:5.1f}".format(BDBR_Y) + '%', "{:5.1f}".format(BDBR_U) + '%', \
              "{:5.1f}".format(BDBR_V)+'%',"{:8.1f}".format(YUVBDBR_avg)+'%', "{:8.1f}".format(SSIM_BDBRY)+'%', \
              "{:8.1f}".format(Delta_time) + '%'
    else:
        print "{:2d}".format(index_num), filename, '-'*(50-len(filename)), "{:5.1f}".format(BDBR_Y) + '%', "{:5.1f}".format(BDBR_U) + '%', \
              "{:5.1f}".format(BDBR_V)+'%',"{:8.1f}".format(YUVBDBR_avg), \
              "{:8.1f}".format(Delta_time) + '%'
    # 两路对比
    if (SSIM == 1):
        oneline = filename +' ' + str(BDBR_Y) + ' ' + str(BDBR_U) + ' ' + str(BDBR_V) + ' ' + str(Delta_YPSNR) + ' ' \
                + str(Delta_UPSNR) + ' ' + str(Delta_VPSNR) + ' ' + str(YUVBDBR_avg) + ' ' + str(SSIM_BDBRY) + ' ' + str(Delta_time) + ' '
    else:
        oneline = filename +' ' + str(BDBR_Y) + ' ' + str(BDBR_U) + ' ' + str(BDBR_V) + ' ' + str(Delta_YPSNR) + ' ' \
            + str(Delta_UPSNR) + ' ' + str(Delta_VPSNR) + ' ' + str(YUVBDBR_avg) + ' ' + str(Delta_time) + ' '
    if (len(sys.argv) > 5):  # 三路对比
        if (SSIM == 1):
            oneline += str(BDBR_Y_2) + ' ' + str(BDBR_U_2) + ' ' + str(BDBR_V_2) + ' ' + str(Delta_YPSNR_2) + ' ' \
                    + str(Delta_UPSNR_2) + ' ' + str(Delta_VPSNR_2) + ' ' + str(YUVBDBR_avg_2) + ' ' + str(SSIM_BDBRY_2) + ' ' + str(Delta_time_2) + '\n'
        else:
            oneline += str(BDBR_Y_2) + ' ' + str(BDBR_U_2) + ' ' + str(BDBR_V_2) + ' ' + str(Delta_YPSNR_2) + ' ' \
                    + str(Delta_UPSNR_2) + ' ' + str(Delta_VPSNR_2) + ' ' + str(YUVBDBR_avg_2) + ' ' + str(Delta_time_2) + '\n'
      
    if (len(sys.argv) > 6): # 四路对比
        if (SSIM == 1):
            oneline += str(BDBR_Y_3) + ' ' + str(BDBR_U_3) + ' ' + str(BDBR_V_3) + ' ' + str(Delta_YPSNR_3) + ' '   \
                    + str(Delta_UPSNR_3) + ' ' + str(Delta_VPSNR_3) + ' ' + str(YUVBDBR_avg_3) + ' ' + str(SSIM_BDBRY_3) + ' ' + str(Delta_time_3) + '\n'
        else:
            oneline += str(BDBR_Y_3) + ' ' + str(BDBR_U_3) + ' ' + str(BDBR_V_3) + ' ' + str(Delta_YPSNR_3) + ' '   \
                    + str(Delta_UPSNR_3) + ' ' + str(Delta_VPSNR_3) + ' ' + str(YUVBDBR_avg_3) + ' ' + str(Delta_time_3) + '\n'

    #print oneline
    return oneline

####################################main 函数入口####################################################
if __name__ == '__main__':
    if(len(sys.argv) < 4):
        print('Usage: auto_data_analysis.py ' + '<anchor outDir SSIM refer1> ' + '[refer2 refer3 ]' + '\n')
        print("For example: auto_data_analysis.py anchor_result ./out 1 refer1_result refer2_result refer3_result")
        print('Notice: <> is necessary, [] is optional')
        exit()
    
    ## 1.命令行参数解析    
    anchor = sys.argv[1]
    outDir = sys.argv[2]
    SSIM_enable = int(sys.argv[3])
    refer1 = sys.argv[4]
    
    ## 2.创建输出目录
    make_all_dir(outDir)

    ## 3.保存数据分析结果的字典  
    origBit_dict   = collections.OrderedDict()  ## key: seq_name value: bitrate
    origYPSNR_dict = collections.OrderedDict()
    origUPSNR_dict = collections.OrderedDict()
    origVPSNR_dict = collections.OrderedDict()
    testBit_dict   = collections.OrderedDict()
    testYPSNR_dict = collections.OrderedDict()
    testUPSNR_dict = collections.OrderedDict()
    testVPSNR_dict = collections.OrderedDict()
    origTime_dict  = collections.OrderedDict()
    testTime_dict  = collections.OrderedDict()
    seqName_dict   = collections.OrderedDict()  ## key: index_num  value: seq_name
    origYSSIM_dict = collections.OrderedDict()
    testYSSIM_dict = collections.OrderedDict()
    if (len(sys.argv) > 5):
        refer2 = sys.argv[5]
        testBit_dict2   = collections.OrderedDict()
        testYPSNR_dict2 = collections.OrderedDict()
        testUPSNR_dict2 = collections.OrderedDict()
        testVPSNR_dict2 = collections.OrderedDict()
        testTime_dict2  = collections.OrderedDict()
        testYSSIM_dict2 = collections.OrderedDict()
    if (len(sys.argv) > 6):
        refer3 = sys.argv[6]
        testBit_dict3   = collections.OrderedDict()
        testYPSNR_dict3 = collections.OrderedDict()
        testUPSNR_dict3 = collections.OrderedDict()
        testVPSNR_dict3 = collections.OrderedDict()
        testTime_dict3  = collections.OrderedDict()
        testYSSIM_dict3 = collections.OrderedDict()

    ## 4.读取anchor数据并提取bitrate, PSNR, SSIM和time信息
    seq_num = shuffle_info(anchor, 1)

    ## 5.读取refer数据并提取bitrate, PSNR, SSIM和time信息
    shuffle_info(refer1, 0, 0)
    if (len(sys.argv) > 5):
        shuffle_info(refer2, 0, 1)
    if (len(sys.argv) > 6):
        shuffle_info(refer3, 0, 2)
    
    anchor_codec = anchor.split('__result')[1].split('_')[1]
    refer1_codec = refer1.split('__result')[1].split('_')[1]
    if (len(sys.argv) > 5):
        refer2_codec = refer2.split('__result')[1].split('_')[1]
    if (len(sys.argv) > 6):
        refer3_codec = refer3.split('__result')[1].split('_')[1]

    ## 6.创建数据分析结果csv文件
    outExcelData = outDir+delimiter+'__result_'+anchor_codec+ '_vs._'+refer1_codec+'_BDBR.csv'
    if (len(sys.argv) > 5):
        outExcelData = outDir+delimiter+'__result_'+anchor_codec+ '_vs._'+refer1_codec+'_vs._'+refer2_codec+'_BDBR.csv'
    if (len(sys.argv) > 6):
        outExcelData = outDir+delimiter+'__result_'+anchor_codec+ '_vs._'+refer1_codec+'_vs._'+refer2_codec+'_vs._'+refer3_codec+'_BDBR.csv'
    create_excel(outExcelData)
    pFile = open(outExcelData, 'w') #创建汇总文件，性能数据
    pFile.write(codecs.BOM_UTF8)
    csv_writer=csv.writer(pFile, dialect='excel')

    #写入输出结果Excel的标题, 默认两路对比
    if (SSIM_enable == 1):
        totaltitle=['video sequence', 'BDBRY(%)', 'BDBRU(%)', 'BDBRV(%)', 'DeltaYPSNR(dB)', 'DeltaUPSNR(dB)', 'DeltaVPSNR(dB)', 'YUV-BDBR(%)', 'SSIM-BDBR(%)','DeltaTime(%)']
    else:
        totaltitle=['video sequence', 'BDBRY(%)', 'BDBRU(%)', 'BDBRV(%)', 'DeltaYPSNR(dB)', 'DeltaUPSNR(dB)', 'DeltaVPSNR(dB)', 'YUV-BDBR(%)','DeltaTime(%)']  
    if (len(sys.argv) > 5):  # 三路对比
        if (SSIM_enable == 1):
            totaltitle=['video sequence', 'BDBRY-1(%)', 'BDBRU-1(%)', 'BDBRV-1(%)', 'DeltaYPSNR-1(dB)', 'DeltaUPSNR-1(dB)', 'DeltaVPSNR-1(dB)', 'YUV-BDBR_1(%)', 'SSIM-BDBR_1(%)', 'DeltaTime_1(%)',\
                                          'BDBRY-2(%)', 'BDBRU-2(%)', 'BDBRV-2(%)', 'DeltaYPSNR-2(dB)', 'DeltaUPSNR-2(dB)', 'DeltaVPSNR-2(dB)', 'YUV-BDBR_2(%)', 'SSIM-BDBR_2(%)', 'DeltaTime_2(%)'  ]
        else:
            totaltitle=['video sequence', 'BDBRY-1(%)', 'BDBRU-1(%)', 'BDBRV-1(%)', 'DeltaYPSNR-1(dB)', 'DeltaUPSNR-1(dB)', 'DeltaVPSNR-1(dB)', 'YUV-BDBR_1(%)', 'DeltaTime_1(%)',\
                                          'BDBRY-2(%)', 'BDBRU-2(%)', 'BDBRV-2(%)', 'DeltaYPSNR-2(dB)', 'DeltaUPSNR-2(dB)', 'DeltaVPSNR-2(dB)', 'YUV-BDBR_2(%)', 'DeltaTime_2(%)'  ]
        
    if (len(sys.argv) > 6): # 四路对比
        if (SSIM_enable == 1):
            totaltitle=['video sequence', 'BDBRY-1(%)', 'BDBRU-1(%)', 'BDBRV-1(%)', 'DeltaYPSNR-1(dB)', 'DeltaUPSNR-1(dB)', 'DeltaVPSNR-1(dB)', 'YUV-BDBR_1(%)', 'SSIM-BDBR_1(%)', 'Delta_time_1(%)', \
                                          'BDBRY-2(%)', 'BDBRU-2(%)', 'BDBRV-2(%)', 'DeltaYPSNR-2(dB)', 'DeltaUPSNR-2(dB)', 'DeltaVPSNR-2(dB)', 'YUV-BDBR_2(%)', 'SSIM-BDBR_2(%)', 'Delta_time_2(%)', \
                                          'BDBRY-3(%)', 'BDBRU-3(%)', 'BDBRV-3(%)', 'DeltaYPSNR-3(dB)', 'DeltaUPSNR-3(dB)', 'DeltaVPSNR-3(dB)', 'YUV-BDBR_3(%)', 'SSIM-BDBR_3(%)', 'Delta_time_3(%)'  ]
        else:
            totaltitle=['video sequence', 'BDBRY-1(%)', 'BDBRU-1(%)', 'BDBRV-1(%)', 'DeltaYPSNR-1(dB)', 'DeltaUPSNR-1(dB)', 'DeltaVPSNR-1(dB)', 'YUV-BDBR_1(%)', 'Delta_time_1(%)', \
                                          'BDBRY-2(%)', 'BDBRU-2(%)', 'BDBRV-2(%)', 'DeltaYPSNR-2(dB)', 'DeltaUPSNR-2(dB)', 'DeltaVPSNR-2(dB)', 'YUV-BDBR_2(%)', 'Delta_time_2(%)', \
                                          'BDBRY-3(%)', 'BDBRU-3(%)', 'BDBRV-3(%)', 'DeltaYPSNR-3(dB)', 'DeltaUPSNR-3(dB)', 'DeltaVPSNR-3(dB)', 'YUV-BDBR_3(%)', 'Delta_time_3(%)'  ]
         
    csv_writer.writerow(totaltitle)
    pFile.close()

    BDBRY_avg       = 0.0
    BDBRU_avg       = 0.0
    BDBRV_avg       = 0.0
    Delta_YPSNR_avg = 0.0
    Delta_UPSNR_avg = 0.0
    Delta_VPSNR_avg = 0.0
    Delta_time_avg  = 0.0 
    YUVBDBR_avg     = 0.0
    SSIMBDBR_avg    = 0.0
    BDBRY_avg_2      = 0.0
    BDBRU_avg_2      = 0.0
    BDBRV_avg_2      = 0.0
    Delta_YPSNR_avg_2 = 0.0
    Delta_UPSNR_avg_2 = 0.0
    Delta_VPSNR_avg_2 = 0.0 
    Delta_time_avg_2  = 0.0
    YUVBDBR_avg_2     = 0.0
    SSIMBDBR_avg_2    = 0.0
    BDBRY_avg_3       = 0.0
    BDBRU_avg_3       = 0.0
    BDBRV_avg_3       = 0.0
    Delta_YPSNR_avg_3 = 0.0
    Delta_UPSNR_avg_3 = 0.0
    Delta_VPSNR_avg_3 = 0.0
    Delta_time_avg_3  = 0.0
    YUVBDBR_avg_3     = 0.0
    SSIMBDBR_avg_3    = 0.0
  
    if (SSIM_enable == 1):
        print"{:2d}".format(0), 'video sequence', '-'*(50-len('video sequence')), "{:5s}".format('Y-BDBR') + ' ', "{:5s}".format('U-BDBR'), \
             "{:5s}".format('V-BDBR')+' ', "{:5s}".format('YUV-BDBR')+' ', "{:5s}".format('SSIM-BDBR')+' ', "{:5s}".format('Delta_time')
    else:
        print"{:2d}".format(0), 'video sequence', '-'*(50-len('video sequence')), "{:5s}".format('Y-BDBR') + ' ', "{:5s}".format('U-BDBR'), \
             "{:5s}".format('V-BDBR')+' ', "{:5s}".format('YUV-BDBR')+' ', "{:5s}".format('Delta_time')
  
    ## 7. 计算BD-rate(piecewise cubic)和BD-rate(cubic)以及Delta_PSNR和Delta_time
    for index_num in range(1, seq_num + 1):        
        filename = seqName_dict[index_num]

        ## 7.1 计算BD-rate, DeltaPSNR, SSIM-BDBR和DeltaTime
        oneline = ComputeBDBR_PSNR_SSIM_Time(index_num, SSIM_enable)
        
        ## 7.2 保存数据
        pFile = open(outExcelData, 'a+')
        pFile.write(codecs.BOM_UTF8)
        csv_writer=csv.writer(pFile, dialect='excel')
        csv_writer.writerow(oneline.split())
        pFile.close()

    ## 8. 输出平均值信息
    BDBRY_avg       = float('%.1f' %(BDBRY_avg   / seq_num))
    BDBRU_avg       = float('%.1f' %(BDBRU_avg   / seq_num))
    BDBRV_avg       = float('%.1f' %(BDBRV_avg   / seq_num))
    Delta_YPSNR_avg = float('%.3f' %(Delta_YPSNR_avg / seq_num))
    Delta_UPSNR_avg = float('%.3f' %(Delta_UPSNR_avg / seq_num))
    Delta_VPSNR_avg = float('%.3f' %(Delta_VPSNR_avg / seq_num))
    Delta_time_avg  = float('%.3f' %(Delta_time_avg  / seq_num))
    YUVBDBR_avg     = float('%.3f' %((6 * BDBRY_avg +  BDBRU_avg + BDBRV_avg) / 8))
    if (SSIM_enable == 1):
        SSIMBDBR_avg    = float('%.3f' %( SSIMBDBR_avg / seq_num))
    if (len(sys.argv) > 5):
        BDBRY_avg_2       = float('%.1f' %(BDBRY_avg_2   / seq_num))
        BDBRU_avg_2       = float('%.1f' %(BDBRU_avg_2   / seq_num))
        BDBRV_avg_2       = float('%.1f' %(BDBRV_avg_2   / seq_num))
        Delta_YPSNR_avg_2 = float('%.3f' %(Delta_YPSNR_avg_2 / seq_num))
        Delta_UPSNR_avg_2 = float('%.3f' %(Delta_UPSNR_avg_2 / seq_num))
        Delta_VPSNR_avg_2 = float('%.3f' %(Delta_VPSNR_avg_2 / seq_num))
        Delta_time_avg_2  = float('%.3f' %(Delta_time_avg_2  / seq_num))
        YUVBDBR_avg_2     = float('%.3f' %((6 * BDBRY_avg_2 +  BDBRU_avg_2 + BDBRV_avg_2) / 8))
        if (SSIM_enable == 1):
            SSIMBDBR_avg_2    = float('%.3f' %( SSIMBDBR_avg_2 / seq_num))
    if (len(sys.argv) > 6):
        BDBRY_avg_3       = float('%.1f' %(BDBRY_avg_3   / seq_num))
        BDBRU_avg_3       = float('%.1f' %(BDBRU_avg_3   / seq_num))
        BDBRV_avg_3       = float('%.1f' %(BDBRV_avg_3   / seq_num))
        Delta_YPSNR_avg_3 = float('%.3f' %(Delta_YPSNR_avg_3 / seq_num))
        Delta_UPSNR_avg_3 = float('%.3f' %(Delta_UPSNR_avg_3 / seq_num))
        Delta_VPSNR_avg_3 = float('%.3f' %(Delta_VPSNR_avg_3 / seq_num))
        Delta_time_avg_3  = float('%.3f' %(Delta_time_avg_3  / seq_num))   
        YUVBDBR_avg_3     = float('%.3f' %((6 * BDBRY_avg_3 +  BDBRU_avg_3 + BDBRV_avg_3) / 8))
        if (SSIM_enable == 1):
            SSIMBDBR_avg_3    = float('%.3f' %( SSIMBDBR_avg_3 / seq_num))
   
    # 写入最后一行数据
    pFile = open(outExcelData, 'a+') #创建汇总文件，追加性能数据
    pFile.write(codecs.BOM_UTF8)
    csv_writer=csv.writer(pFile, dialect='excel')

    if (SSIM_enable==1):
        average_data = 'Average:\n' + str(BDBRY_avg) + ' ' + str(BDBRU_avg) + ' ' + str(BDBRV_avg) + ' ' + str(Delta_YPSNR_avg)  \
                        + ' ' + str(Delta_UPSNR_avg) + ' ' + str(Delta_VPSNR_avg) + ' ' + str(YUVBDBR_avg) + ' ' + str(SSIMBDBR_avg) + ' ' + str(Delta_time_avg) 
        average_data_show = 'Summary:\n' + anchor_codec + ' vs. ' + refer1_codec + ': ' + '{:8s}'.format(str(BDBRY_avg)) \
                        + '{:8s}'.format(str(BDBRU_avg)) + '{:8s}'.format(str(BDBRV_avg))         \
                        + '{:10s}'.format(str(YUVBDBR_avg)) + '{:8s}'.format(str(SSIMBDBR_avg)) + '{:8s}'.format(str(Delta_time_avg))
    else:
        average_data = 'Average:\n' + str(BDBRY_avg) + ' ' + str(BDBRU_avg) + ' ' + str(BDBRV_avg) + ' ' + str(Delta_YPSNR_avg)  \
                        + ' ' + str(Delta_UPSNR_avg) + ' ' + str(Delta_VPSNR_avg) + ' ' + str(YUVBDBR_avg) + ' ' + str(Delta_time_avg) 
        average_data_show = 'Summary:\n' + anchor_codec + ' vs. ' + refer1_codec + ': ' + '{:8s}'.format(str(BDBRY_avg)) \
                        + '{:8s}'.format(str(BDBRU_avg)) + '{:8s}'.format(str(BDBRV_avg))         \
                        + '{:10s}'.format(str(YUVBDBR_avg)) + '{:8s}'.format(str(Delta_time_avg))
    if (len(sys.argv) > 5):
        if (SSIM_enable == 1):
            average_data += '\n' + str(BDBRY_avg_2) + ' ' + str(BDBRU_avg_2) + ' ' + str(BDBRV_avg_2) + ' ' + str(Delta_YPSNR_avg_2)     \
                         + ' '+str(Delta_UPSNR_avg_2) + ' ' + str(Delta_VPSNR_avg_2) + ' ' + str(YUVBDBR_avg_2) + ' ' + str(SSIMBDBR_avg_2) + ' ' + str(Delta_time_avg_2)  
            average_data_show +=  '\n' + anchor_codec + ' vs.' + ' '*(6-len(refer2_codec)) +  refer2_codec + ': ' \
                            + '{:8s}'.format(str(BDBRY_avg_2)) + '{:8s}'.format(str(BDBRU_avg_2)) + '{:8s}'.format(str(BDBRV_avg_2)) \
                            + '{:10s}'.format(str(YUVBDBR_avg_2)) + '{:8s}'.format(str(SSIMBDBR_avg_2)) + '{:8s}'.format(str(Delta_time_avg_2))  
        else:
            average_data += '\n' + str(BDBRY_avg_2) + ' ' + str(BDBRU_avg_2) + ' ' + str(BDBRV_avg_2) + ' ' + str(Delta_YPSNR_avg_2)     \
                         + ' '+str(Delta_UPSNR_avg_2) + ' ' + str(Delta_VPSNR_avg_2) + ' ' + str(YUVBDBR_avg_2) + ' ' + str(Delta_time_avg_2)  
            average_data_show +=  '\n' + anchor_codec + ' vs.' + ' '*(6-len(refer2_codec)) +  refer2_codec + ': '                    \
                            + '{:8s}'.format(str(BDBRY_avg_2)) + '{:8s}'.format(str(BDBRU_avg_2)) + '{:8s}'.format(str(BDBRV_avg_2)) \
                            + '{:10s}'.format(str(YUVBDBR_avg_2)) + '{:8s}'.format(str(Delta_time_avg_2))  
    if (len(sys.argv) > 6):
        if (SSIM_enable == 1):
            average_data += '\n' + str(BDBRY_avg_3) + ' ' + str(BDBRU_avg_3) + ' ' + str(BDBRV_avg_3) + ' ' + str(Delta_YPSNR_avg_3) + ' '   \
                         + str(Delta_UPSNR_avg_3) + ' ' + str(Delta_VPSNR_avg_3) + ' ' + str(YUVBDBR_avg_3) + ' ' + str(SSIMBDBR_avg_3) + ' ' + str(Delta_time_avg_3)
            average_data_show += '\n' + anchor_codec + ' vs.' + ' '*(6-len(refer3_codec))+ refer3_codec + ': '                   \
                        + '{:8s}'.format(str(BDBRY_avg_3)) + '{:8s}'.format(str(BDBRU_avg_3)) + '{:8s}'.format(str(BDBRV_avg_3)) \
                        + '{:10s}'.format(str(YUVBDBR_avg_3)) + '{:8s}'.format(str(SSIMBDBR_avg_3)) + '{:8s}'.format(str(Delta_time_avg_3))                                               
        else:
            average_data += '\n' + str(BDBRY_avg_3) + ' ' + str(BDBRU_avg_3) + ' ' + str(BDBRV_avg_3) + ' ' + str(Delta_YPSNR_avg_3) + ' '   \
                         + str(Delta_UPSNR_avg_3) + ' ' + str(Delta_VPSNR_avg_3) + ' ' + str(YUVBDBR_avg_3)  + ' ' + str(Delta_time_avg_3)
            average_data_show += '\n' + anchor_codec + ' vs.' + ' '*(6-len(refer3_codec))+ refer3_codec + ': '                  \
                        + '{:8s}'.format(str(BDBRY_avg_3)) + '{:8s}'.format(str(BDBRU_avg_3)) + '{:8s}'.format(str(BDBRV_avg_3))\
                        + '{:10s}'.format(str(YUVBDBR_avg_3)) + '{:8s}'.format(str(Delta_time_avg_3))                                               
    print '\n', average_data_show
    csv_writer.writerow(average_data.split())
    pFile.close()

    ## 9. 将csv文件转换成excel文件,此处为了在一个表格里面绘制率失真曲线图
    analysis_file = outDir+delimiter+'analysis_result_'+anchor_codec+'_vs._'+refer1_codec+'.xlsx'
    if (len(sys.argv) > 5):
        analysis_file = outDir+delimiter+'analysis_result_'+anchor_codec+ '_vs._'+refer1_codec+'_vs._'+refer2_codec+'.xlsx'
    if (len(sys.argv) > 6):
        analysis_file = outDir+delimiter+'analysis_result_'+anchor_codec+ '_vs._'+refer1_codec+'_vs._'+refer2_codec+'_vs._'+refer3_codec+'.xlsx'

    writer    = pd.ExcelWriter(analysis_file)

    csv_file1 = pd.read_csv(outExcelData, encoding='utf-8')
    csv_file1.to_excel(writer, sheet_name='result')

    csv_file2 = pd.read_csv(anchor, encoding='utf-8')
    csv_file2.to_excel(writer, sheet_name='anchor_data')

    csv_file3 = pd.read_csv(refer1, encoding='utf-8')
    csv_file3.to_excel(writer, sheet_name='refer1_data')
    if (len(sys.argv) > 5):
        csv_file4 = pd.read_csv(refer2, encoding='utf-8')
        csv_file4.to_excel(writer, sheet_name='refer2_data') 
    if (len(sys.argv) > 6):
        csv_file5 = pd.read_csv(refer3, encoding='utf-8')
        csv_file5.to_excel(writer, sheet_name='refer3_data') 
    writer.save()

    wb = load_workbook(analysis_file)
    #print wb.get_sheet_names()
    sheet_result = wb.get_sheet_by_name('result')       # 获得当前正在显示的sheet
    sheet_anchor = wb.get_sheet_by_name('anchor_data')  # 获得当前正在显示的sheet
    sheet_refer  = wb.get_sheet_by_name('refer1_data')  # 获得当前正在显示的sheet  
    if (len(sys.argv) > 5):
        sheet_refer2  = wb.get_sheet_by_name('refer2_data')  # 获得当前正在显示的sheet        
    if (len(sys.argv) > 6):
        sheet_refer3  = wb.get_sheet_by_name('refer3_data')  # 获得当前正在显示的sheet   

    excelCurrRow = 2
    dataVerStep  = 4
    ## 10. 绘制率失真曲线图
    for index_num in range(1, seq_num + 1):
        filename = seqName_dict[index_num] 
        line = openpyxl.chart.ScatterChart()
        line.title = filename                   #图表标题
        line.x_axis.title = 'Bitrate (kbps)'    #y轴标题
        line.y_axis.title = 'Y-PSNR (dB)'       #x轴标题

        line.y_axis.scaling.min = int(min(origYPSNR_dict[filename][3], testYPSNR_dict[filename][3])) - 2  # y轴的最小值
        if (len(sys.argv) > 5):
            line.y_axis.scaling.min = int(min(origYPSNR_dict[filename][3], testYPSNR_dict[filename][3],
                                          testYPSNR_dict2[filename][3])) - 5  # y轴的最小值
        if (len(sys.argv) > 6):
            line.y_axis.scaling.min = int(min(origYPSNR_dict[filename][3],  testYPSNR_dict[filename][3],
                                              testYPSNR_dict2[filename][3], testYPSNR_dict3[filename][3])) - 2  # y轴的最小值
            #line.x_axis.scaling.max = int(max(origYPSNR_dict[filename][0] , testYPSNR_dict[filename][0],
            #                                  testYPSNR_dict2[filename][0], testYPSNR_dict3[filename][0])) + 2  # y轴的最大值
        
        oriXdata = Reference(sheet_anchor, min_col=4, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
        oriYdata = Reference(sheet_anchor, min_col=5, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
        series = Series(oriYdata, oriXdata, title=anchor_codec)
        series.marker.symbol = 'circle'
        series.smooth = True
        line.series.append(series)

        testXdata = Reference(sheet_refer, min_col=4, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
        testYdata = Reference(sheet_refer, min_col=5, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
        series = Series(testYdata, testXdata, title=refer1_codec)
        series.marker.symbol = 'circle'
        series.smooth = True
        line.series.append(series)

        if (len(sys.argv) > 5):
            testXdata = Reference(sheet_refer2, min_col=4, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
            testYdata = Reference(sheet_refer2, min_col=5, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
            series = Series(testYdata, testXdata, title=refer2_codec)
            series.marker.symbol = 'circle'
            series.smooth = True
            line.series.append(series)

        if (len(sys.argv) > 6):
            testXdata = Reference(sheet_refer3, min_col=4, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
            testYdata = Reference(sheet_refer3, min_col=5, min_row=excelCurrRow + (index_num-1) * dataVerStep, max_row=excelCurrRow + (index_num-1) * dataVerStep+3)
            series = Series(testYdata, testXdata, title=refer3_codec)
            series.marker.symbol = 'circle'
            series.smooth = True
            line.series.append(series)

        # 两路对比
        if (index_num-1) % 2 == 0:
            chartColumn = 'L'
        else:
            chartColumn = 'U'
        # 三路对比
        if (len(sys.argv) > 5):
            if (index_num-1) % 2 == 0:
                chartColumn = 'U'
            else:
                chartColumn = 'AD'
        # 四路对比
        if (len(sys.argv) > 6):
            if (index_num-1) % 2 == 0:
                chartColumn = 'AD'
            else:
                chartColumn = 'AM'          
        sheet_result.add_chart(line, chartColumn + str(excelCurrRow + (index_num-1) * (dataVerStep+4)+2))       
    # 保存分析结果文件
    wb.save(analysis_file)
    
    ret = 0
    if(ret != -1):
        print("---------Process finished!---------")
        os._exit(0)